from datetime import date, datetime
from collections import defaultdict
from decimal import Decimal

from django.db.models import Sum, Count, Q
from django.http import HttpResponse
from django.utils.timezone import now

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch, mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Frame, PageTemplate, BaseDocTemplate
)
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

from .permissions import IsHODOrStorekeeper

from stock_register.models import StockRegister, ChemicalItem, ApparatusItem
from stock_request.models import (
    StockRequest, StockRequestChemicalItem, StockRequestApparatusItem,
    IssueRegister, IssueChemicals
)
from damaged_entry.models import DamagedEntry, DamagedItem
from inventory.models import AvailableChemical, AvailableApparatus


def get_academic_year_range(year=None):
    if year is None:
        today = date.today()
        if today.month >= 6:
            year = today.year
        else:
            year = today.year - 1
    else:
        year = int(year)
    start_date = date(year, 6, 1)
    end_date = date(year + 1, 5, 31)
    return year, year + 1, start_date, end_date


class YearEndReportView(APIView):
    permission_classes = [IsHODOrStorekeeper]

    def get(self, request):
        year_param = request.query_params.get('year')
        start_year, end_year, start_date, end_date = get_academic_year_range(year_param)

        data = self.build_report(start_year, end_year, start_date, end_date)
        return Response(data)

    def build_report(self, start_year, end_year, start_date, end_date):
        stock_registers = StockRegister.objects.filter(date__range=[start_date, end_date])
        stock_register_ids = stock_registers.values_list('id', flat=True)

        issue_registers = IssueRegister.objects.filter(date__range=[start_date, end_date])
        ir_ids = issue_registers.values_list('ir_id', flat=True)

        damaged_entries = DamagedEntry.objects.filter(date__range=[start_date, end_date])
        damaged_entry_ids = damaged_entries.values_list('id', flat=True)

        stock_requests = StockRequest.objects.filter(
            created_at__date__range=[start_date, end_date]
        )

        # --- Chemical Items ---
        chem_items = ChemicalItem.objects.filter(stock_register_id__in=stock_register_ids)
        chem_items_list = list(chem_items)
        chem_by_name = defaultdict(lambda: {'total_qty': Decimal('0'), 'total_cost': Decimal('0'), 'count': 0, 'unit': 'ml'})
        for item in chem_items_list:
            name = item.chemical_name
            chem_by_name[name]['total_qty'] += item.total_quantity
            chem_by_name[name]['total_cost'] += item.total_quantity * item.rate
            chem_by_name[name]['count'] += 1
            chem_by_name[name]['unit'] = item.unit

        # --- Apparatus Items ---
        app_items = ApparatusItem.objects.filter(stock_register_id__in=stock_register_ids)
        app_items_list = list(app_items)
        app_by_name = defaultdict(lambda: {'total_qty': 0, 'total_cost': Decimal('0'), 'count': 0})
        for item in app_items_list:
            name = item.apparatus_name
            app_by_name[name]['total_qty'] += item.quantity_pieces
            app_by_name[name]['total_cost'] += item.quantity_pieces * item.rate
            app_by_name[name]['count'] += 1

        # --- Monthly Purchase Trend ---
        monthly_map = defaultdict(lambda: {
            'chemicals_cost': Decimal('0'), 'apparatus_cost': Decimal('0'),
            'chemicals_quantity': Decimal('0'), 'apparatus_quantity': 0
        })
        for sr in stock_registers:
            month_key = sr.date.strftime('%b %Y')
            for ci in ChemicalItem.objects.filter(stock_register=sr):
                monthly_map[month_key]['chemicals_cost'] += ci.total_quantity * ci.rate
                monthly_map[month_key]['chemicals_quantity'] += ci.total_quantity
            for ai in ApparatusItem.objects.filter(stock_register=sr):
                monthly_map[month_key]['apparatus_cost'] += ai.quantity_pieces * ai.rate
                monthly_map[month_key]['apparatus_quantity'] += ai.quantity_pieces

        # Build sorted monthly data
        all_months = []
        for m in range(6, 13):
            all_months.append(date(start_year, m, 1))
        for m in range(1, 6):
            all_months.append(date(end_year, m, 1))

        monthly_purchase_trend = []
        for m in all_months:
            key = m.strftime('%b %Y')
            entry = monthly_map.get(key, {
                'chemicals_cost': Decimal('0'), 'apparatus_cost': Decimal('0'),
                'chemicals_quantity': Decimal('0'), 'apparatus_quantity': 0
            })
            monthly_purchase_trend.append({
                'month': key,
                'chemicals_cost': float(entry['chemicals_cost']),
                'apparatus_cost': float(entry['apparatus_cost']),
                'total_cost': float(entry['chemicals_cost'] + entry['apparatus_cost']),
                'chemicals_quantity': float(entry['chemicals_quantity']),
                'apparatus_quantity': entry['apparatus_quantity'],
            })

        # --- Usage / Issue Chemicals ---
        issue_chems = IssueChemicals.objects.filter(ir_id__in=ir_ids)
        chem_usage_map = defaultdict(lambda: {'total_used': Decimal('0'), 'unit': 'ml', 'times': 0})
        for ic in issue_chems:
            name = ic.chemical_name
            actual = ic.actual_usage if ic.actual_usage is not None else ic.issued_quantity
            chem_usage_map[name]['total_used'] += actual
            chem_usage_map[name]['times'] += 1

        # Also get from StockRequestChemicalItem actual_used_quantity
        sr_chem_items = StockRequestChemicalItem.objects.filter(
            stock_request__in=stock_requests,
            actual_used_quantity__isnull=False
        )
        for item in sr_chem_items:
            name = item.chemical_name
            chem_usage_map[name]['total_used'] += item.actual_used_quantity
            chem_usage_map[name]['unit'] = item.unit
            chem_usage_map[name]['times'] += 1

        top_used = sorted(
            [{'name': k, 'total_used': float(v['total_used']), 'unit': v['unit'], 'times_requested': v['times']}
             for k, v in chem_usage_map.items()],
            key=lambda x: x['total_used'], reverse=True
        )[:10]

        # --- Purchases ---
        purchases_chemicals = [
            {'name': k, 'total_quantity': float(v['total_qty']), 'unit': v['unit'],
             'total_cost': float(v['total_cost']), 'purchase_count': v['count']}
            for k, v in chem_by_name.items()
        ]
        purchases_apparatus = [
            {'name': k, 'total_quantity': v['total_qty'], 'unit': 'nos',
             'total_cost': float(v['total_cost']), 'purchase_count': v['count']}
            for k, v in app_by_name.items()
        ]

        # --- Usage by Class ---
        completed_requests = StockRequest.objects.filter(
            status='completed',
            created_at__date__range=[start_date, end_date]
        )
        usage_by_class_map = defaultdict(lambda: {'requests': 0, 'chemicals_used': Decimal('0')})

        for sr in completed_requests:
            cls = sr.class_name
            usage_by_class_map[cls]['requests'] += 1
            irs = IssueRegister.objects.filter(stock_request_db_id=sr.id)
            for ir in irs:
                for ic in IssueChemicals.objects.filter(ir=ir):
                    actual = ic.actual_usage if ic.actual_usage is not None else ic.issued_quantity
                    usage_by_class_map[cls]['chemicals_used'] += actual

        for sr in completed_requests:
            cls = sr.class_name
            for item in StockRequestChemicalItem.objects.filter(
                stock_request=sr, actual_used_quantity__isnull=False
            ):
                usage_by_class_map[cls]['chemicals_used'] += item.actual_used_quantity

        usage_by_class = [
            {'class_name': k, 'total_requests': v['requests'],
             'total_chemicals_used': float(v['chemicals_used'])}
            for k, v in usage_by_class_map.items()
        ]

        # --- Damage Summary ---
        damage_items = DamagedItem.objects.filter(damaged_entry_id__in=damaged_entry_ids)
        # DamagedItem only has apparatus_name, not chemical_name
        damage_app_map = defaultdict(lambda: {'total_qty': 0, 'count': 0})
        for di in damage_items:
            name = di.apparatus_name
            damage_app_map[name]['total_qty'] += di.quantity
            damage_app_map[name]['count'] += 1

        damage_chem = []
        damage_app = [
            {'name': k, 'total_quantity': v['total_qty'], 'unit': 'nos',
             'incident_count': v['count']}
            for k, v in damage_app_map.items()
        ]

        # --- Summary Totals ---
        total_chemicals_purchased = len(chem_by_name)
        total_apparatus_purchased = len(app_by_name)
        total_chemicals_used = float(sum(v['total_used'] for v in chem_usage_map.values()))
        total_apparatus_used = sum(v['total_qty'] for v in app_by_name.values())
        total_damaged_chemicals = sum(v['total_qty'] for v in damage_chem) if damage_chem else 0
        total_damaged_apparatus = sum(v['total_qty'] for v in damage_app_map.values())
        total_spend = float(sum(
            v['total_cost'] for v in chem_by_name.values()
        ) + sum(
            v['total_cost'] for v in app_by_name.values()
        ))

        summary = {
            'total_spend': round(total_spend, 2),
            'total_chemicals_purchased': total_chemicals_purchased,
            'total_apparatus_purchased': total_apparatus_purchased,
            'total_chemicals_used': round(total_chemicals_used, 2),
            'total_apparatus_used': total_apparatus_used,
            'total_damaged_chemicals': total_damaged_chemicals,
            'total_damaged_apparatus': total_damaged_apparatus,
        }

        # --- Data Integrity Warnings ---
        data_warnings = []

        # --- Current Stock (low stock) ---
        avail_chems = AvailableChemical.objects.all()
        low_chems = []
        for ac in avail_chems:
            rl = float(ac.reorder_level) if ac.reorder_level is not None else 0
            raw_qty = float(ac.quantity)
            current_qty = max(0, raw_qty)
            if raw_qty < 0:
                data_warnings.append(f"Negative stock detected for chemical '{ac.chemical_name}': {raw_qty}")
            if current_qty <= rl:
                low_chems.append({
                    'name': ac.chemical_name,
                    'current_quantity': current_qty,
                    'unit': ac.unit,
                    'reorder_level': float(rl),
                    'recommended_purchase': float(max(0, rl - current_qty)),
                })

        avail_apps = AvailableApparatus.objects.all()
        low_apps = []
        for aa in avail_apps:
            rl = aa.reorder_level if aa.reorder_level is not None else 0
            raw_qty = aa.available_quantity_pieces
            current_qty = max(0, raw_qty)
            if raw_qty < 0:
                data_warnings.append(f"Negative stock detected for apparatus '{aa.apparatus_name}': {raw_qty}")
            if current_qty <= rl:
                low_apps.append({
                    'name': aa.apparatus_name,
                    'current_quantity': current_qty,
                    'unit': 'nos',
                    'reorder_level': rl,
                    'recommended_purchase': max(0, rl - current_qty),
                })

        current_stock = {
            'low_stock_chemicals': low_chems,
            'low_stock_apparatus': low_apps,
        }

        # --- Restock Recommendations ---
        restock_recs = []
        # For chemicals
        all_chem_names = set()
        for item in chem_items_list:
            all_chem_names.add(item.chemical_name)
        for item in list(chem_usage_map.keys()):
            all_chem_names.add(item)

        for name in all_chem_names:
            total_used = chem_usage_map.get(name, {}).get('total_used', Decimal('0'))
            avg_monthly = float(total_used) / 12 if total_used > 0 else 0
            projected = avg_monthly * 12
            current_qty = 0
            unit = chem_usage_map.get(name, {}).get('unit', 'ml')
            try:
                ac = AvailableChemical.objects.get(chemical_name__iexact=name)
                raw_qty = float(ac.quantity)
                if raw_qty < 0:
                    data_warnings.append(f"Negative stock detected for chemical '{name}': {raw_qty}")
                current_qty = max(0, raw_qty)
                unit = ac.unit
            except AvailableChemical.DoesNotExist:
                pass
            recommended = max(0, projected - current_qty)
            if recommended > 0:
                restock_recs.append({
                    'name': name,
                    'type': 'chemical',
                    'avg_monthly_usage': round(avg_monthly, 2),
                    'projected_annual_need': round(projected, 2),
                    'current_stock': round(current_qty, 2),
                    'recommended_purchase': round(recommended, 2),
                    'unit': unit,
                })

        # For apparatus (usage data is limited, use purchase history)
        app_names_purchased = set()
        for item in app_items_list:
            app_names_purchased.add(item.apparatus_name)
        for name in app_names_purchased:
            total_purchased = app_by_name.get(name, {}).get('total_qty', 0)
            avg_monthly = total_purchased / 12 if total_purchased > 0 else 0
            projected = avg_monthly * 12
            current_qty = 0
            try:
                aa = AvailableApparatus.objects.get(apparatus_name__iexact=name)
                raw_qty = aa.available_quantity_pieces
                if raw_qty < 0:
                    data_warnings.append(f"Negative stock detected for apparatus '{name}': {raw_qty}")
                current_qty = max(0, raw_qty)
            except AvailableApparatus.DoesNotExist:
                pass
            recommended = max(0, projected - current_qty)
            if recommended > 0:
                restock_recs.append({
                    'name': name,
                    'type': 'apparatus',
                    'avg_monthly_usage': round(avg_monthly, 2),
                    'projected_annual_need': round(projected, 2),
                    'current_stock': current_qty,
                    'recommended_purchase': recommended,
                    'unit': 'nos',
                })

        return {
            'academic_year': f"{start_year}-{end_year}",
            'date_range': {
                'start': start_date.isoformat(),
                'end': end_date.isoformat(),
            },
            'summary': summary,
            'monthly_purchase_trend': monthly_purchase_trend,
            'top_used_chemicals': top_used,
            'purchases': {
                'chemicals': purchases_chemicals,
                'apparatus': purchases_apparatus,
            },
            'usage_by_class': usage_by_class,
            'damage_summary': {
                'chemicals': damage_chem,
                'apparatus': damage_app,
            },
            'current_stock': current_stock,
            'restock_recommendations': restock_recs,
            'data_integrity_warnings': data_warnings,
        }


class YearEndPDFDownloadView(APIView):
    permission_classes = [IsHODOrStorekeeper]

    def get(self, request):
        year_param = request.query_params.get('year')
        start_year, end_year, start_date, end_date = get_academic_year_range(year_param)
        report_data = YearEndReportView().build_report(start_year, end_year, start_date, end_date)

        response = HttpResponse(content_type='application/pdf')
        filename = f"year_end_report_{start_year}_{end_year}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        doc = SimpleDocTemplate(
            response,
            pagesize=A4,
            topMargin=20*mm,
            bottomMargin=20*mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Title'],
            fontSize=18, textColor=colors.HexColor('#1E3A8A'),
            spaceAfter=6, alignment=TA_CENTER
        )
        subtitle_style = ParagraphStyle(
            'CustomSubtitle', parent=styles['Normal'],
            fontSize=12, textColor=colors.HexColor('#1E3A8A'),
            spaceAfter=4, alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading', parent=styles['Heading2'],
            fontSize=14, textColor=colors.HexColor('#1E3A8A'),
            spaceBefore=12, spaceAfter=6
        )
        normal_style = ParagraphStyle(
            'CustomNormal', parent=styles['Normal'],
            fontSize=10, spaceAfter=4
        )

        elements = []

        # Cover page
        elements.append(Spacer(1, 40*mm))
        elements.append(Paragraph("Guru Nanak College", title_style))
        elements.append(Paragraph("(PG and Research Programme)", subtitle_style))
        elements.append(Paragraph("Department of Chemistry", subtitle_style))
        elements.append(Spacer(1, 10*mm))
        elements.append(Paragraph("Year-End Audit Report", ParagraphStyle(
            'ReportTitle', parent=styles['Title'],
            fontSize=22, textColor=colors.HexColor('#1E3A8A'),
            spaceAfter=12, alignment=TA_CENTER
        )))
        elements.append(Paragraph(
            f"Academic Year {report_data['academic_year']}", subtitle_style
        ))
        elements.append(Paragraph(
            f"Generated on: {now().strftime('%d %B %Y')}", subtitle_style
        ))

        elements.append(Spacer(1, 15*mm))
        s = report_data['summary']
        summary_data = [
            ['Total Spend', 'Chemicals Purchased', 'Damaged Items'],
            [f"Rs. {s['total_spend']:,.2f}", str(s['total_chemicals_purchased']),
             str(s['total_damaged_chemicals'] + s['total_damaged_apparatus'])],
        ]
        summary_table = Table(summary_data, colWidths=[60*mm, 60*mm, 60*mm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#F8FAFC')),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.HexColor('#F8FAFC'), colors.white]),
        ]))
        elements.append(summary_table)
        elements.append(PageBreak())

        # Page 2 - Purchases
        elements.append(Paragraph("Chemicals Purchased", heading_style))
        chem_purchases = report_data['purchases']['chemicals']
        if chem_purchases:
            chem_table_data = [['Name', 'Quantity', 'Unit', 'Total Cost (Rs.)']]
            for c in chem_purchases:
                chem_table_data.append([
                    c['name'], str(c['total_quantity']), c['unit'],
                    f"{c['total_cost']:,.2f}"
                ])
            self._add_table(elements, chem_table_data)
        else:
            elements.append(Paragraph("No chemicals purchased in this period.", normal_style))

        elements.append(Paragraph("Apparatus Purchased", heading_style))
        app_purchases = report_data['purchases']['apparatus']
        if app_purchases:
            app_table_data = [['Name', 'Quantity', 'Unit', 'Total Cost (Rs.)']]
            for a in app_purchases:
                app_table_data.append([
                    a['name'], str(a['total_quantity']), a['unit'],
                    f"{a['total_cost']:,.2f}"
                ])
            self._add_table(elements, app_table_data)
        else:
            elements.append(Paragraph("No apparatus purchased in this period.", normal_style))
        elements.append(PageBreak())

        # Page 3 - Usage
        elements.append(Paragraph("Top Used Chemicals", heading_style))
        top_used = report_data['top_used_chemicals']
        if top_used:
            usage_table_data = [['Chemical Name', 'Total Used', 'Unit', 'Times Requested']]
            for u in top_used:
                usage_table_data.append([
                    u['name'], str(u['total_used']), u['unit'], str(u['times_requested'])
                ])
            self._add_table(elements, usage_table_data)
        else:
            elements.append(Paragraph("No usage data found.", normal_style))

        elements.append(Paragraph("Usage by Class", heading_style))
        usage_cls = report_data['usage_by_class']
        if usage_cls:
            cls_table_data = [['Class Name', 'Total Requests', 'Chemicals Used']]
            for u in usage_cls:
                cls_table_data.append([
                    u['class_name'], str(u['total_requests']), str(u['total_chemicals_used'])
                ])
            self._add_table(elements, cls_table_data)
        else:
            elements.append(Paragraph("No class usage data found.", normal_style))
        elements.append(PageBreak())

        # Page 4 - Damage Log
        elements.append(Paragraph("Damage Log - Chemicals", heading_style))
        dmg_chem = report_data['damage_summary']['chemicals']
        if dmg_chem:
            dmg_chem_data = [['Chemical Name', 'Quantity', 'Unit', 'Incidents']]
            for d in dmg_chem:
                dmg_chem_data.append([
                    d['name'], str(d['total_quantity']), d['unit'], str(d['incident_count'])
                ])
            self._add_table(elements, dmg_chem_data)
        else:
            elements.append(Paragraph("No chemical damage records.", normal_style))

        elements.append(Paragraph("Damage Log - Apparatus", heading_style))
        dmg_app = report_data['damage_summary']['apparatus']
        if dmg_app:
            dmg_app_data = [['Apparatus Name', 'Quantity', 'Unit', 'Incidents']]
            for d in dmg_app:
                dmg_app_data.append([
                    d['name'], str(d['total_quantity']), d['unit'], str(d['incident_count'])
                ])
            self._add_table(elements, dmg_app_data)
        else:
            elements.append(Paragraph("No apparatus damage records.", normal_style))
        elements.append(PageBreak())

        # Page 5 - Restock Recommendations
        elements.append(Paragraph("Restock Recommendations", heading_style))
        recs = report_data['restock_recommendations']
        if recs:
            rec_table_data = [['Name', 'Type', 'Avg Monthly', 'Annual Need',
                               'Current Stock', 'Recommended', 'Unit']]
            for r in recs:
                rec_table_data.append([
                    r['name'], r['type'], str(r['avg_monthly_usage']),
                    str(r['projected_annual_need']), str(r['current_stock']),
                    str(r['recommended_purchase']), r['unit']
                ])
            self._add_table(elements, rec_table_data)
        else:
            elements.append(Paragraph("No restock recommendations.", normal_style))

        # Build
        doc.build(elements)
        return response

    def _add_table(self, elements, data):
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1),
             [colors.HexColor('#F8FAFC'), colors.white]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 6*mm))


class YearEndExcelDownloadView(APIView):
    permission_classes = [IsHODOrStorekeeper]

    def get(self, request):
        year_param = request.query_params.get('year')
        start_year, end_year, start_date, end_date = get_academic_year_range(year_param)
        report_data = YearEndReportView().build_report(start_year, end_year, start_date, end_date)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"year_end_report_{start_year}_{end_year}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb = openpyxl.Workbook()

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        alt_fill = PatternFill(start_color='F0F4FF', end_color='F0F4FF', fill_type='solid')
        center_align = Alignment(horizontal='center')

        # Sheet 1 - Summary
        ws1 = wb.active
        ws1.title = 'Summary'
        ws1.sheet_properties.tabColor = '1E3A8A'
        s = report_data['summary']
        summary_rows = [
            ('Academic Year', report_data['academic_year']),
            ('Generated On', now().strftime('%d %B %Y')),
            ('Total Spend (Rs.)', s['total_spend']),
            ('Total Chemicals Purchased', s['total_chemicals_purchased']),
            ('Total Apparatus Purchased', s['total_apparatus_purchased']),
            ('Total Chemicals Used', s['total_chemicals_used']),
            ('Total Damaged Items', s['total_damaged_chemicals'] + s['total_damaged_apparatus']),
        ]
        for i, (field, value) in enumerate(summary_rows, 1):
            ws1.cell(row=i, column=1, value=field).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=value)
        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 20

        # Sheet 2 - Purchases
        ws2 = wb.create_sheet('Purchases')
        ws2.sheet_properties.tabColor = '10B981'
        headers = ['Type', 'Name', 'Quantity', 'Unit', 'Cost (Rs.)', 'Purchase Count']
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        row = 2
        for c in report_data['purchases']['chemicals']:
            ws2.cell(row=row, column=1, value='Chemical')
            ws2.cell(row=row, column=2, value=c['name'])
            ws2.cell(row=row, column=3, value=c['total_quantity']).number_format = '#,##0.00'
            ws2.cell(row=row, column=4, value=c['unit'])
            ws2.cell(row=row, column=5, value=c['total_cost']).number_format = '#,##0.00'
            ws2.cell(row=row, column=6, value=c['purchase_count'])
            if row % 2 == 0:
                for col in range(1, 7):
                    ws2.cell(row=row, column=col).fill = alt_fill
            row += 1
        for a in report_data['purchases']['apparatus']:
            ws2.cell(row=row, column=1, value='Apparatus')
            ws2.cell(row=row, column=2, value=a['name'])
            ws2.cell(row=row, column=3, value=a['total_quantity'])
            ws2.cell(row=row, column=4, value=a['unit'])
            ws2.cell(row=row, column=5, value=a['total_cost']).number_format = '#,##0.00'
            ws2.cell(row=row, column=6, value=a['purchase_count'])
            if row % 2 == 0:
                for col in range(1, 7):
                    ws2.cell(row=row, column=col).fill = alt_fill
            row += 1
        self._auto_width(ws2, 6)

        # Sheet 3 - Usage
        ws3 = wb.create_sheet('Usage')
        ws3.sheet_properties.tabColor = '3B82F6'
        headers = ['Chemical Name', 'Total Used', 'Unit', 'Times Requested', 'Class Name']
        for col, h in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        row = 2
        for u in report_data['top_used_chemicals']:
            ws3.cell(row=row, column=1, value=u['name'])
            ws3.cell(row=row, column=2, value=u['total_used']).number_format = '#,##0.00'
            ws3.cell(row=row, column=3, value=u['unit'])
            ws3.cell(row=row, column=4, value=u['times_requested'])
            ws3.cell(row=row, column=5, value='')
            if row % 2 == 0:
                for col in range(1, 6):
                    ws3.cell(row=row, column=col).fill = alt_fill
            row += 1
        for u in report_data['usage_by_class']:
            ws3.cell(row=row, column=1, value='')
            ws3.cell(row=row, column=2, value=u['total_chemicals_used']).number_format = '#,##0.00'
            ws3.cell(row=row, column=3, value='')
            ws3.cell(row=row, column=4, value=u['total_requests'])
            ws3.cell(row=row, column=5, value=u['class_name'])
            if row % 2 == 0:
                for col in range(1, 6):
                    ws3.cell(row=row, column=col).fill = alt_fill
            row += 1
        self._auto_width(ws3, 5)

        # Sheet 4 - Damage Log
        ws4 = wb.create_sheet('Damage Log')
        ws4.sheet_properties.tabColor = 'EF4444'
        headers = ['Type', 'Name', 'Quantity', 'Unit', 'Reason', 'Incident Count']
        for col, h in enumerate(headers, 1):
            cell = ws4.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        row = 2
        for d in report_data['damage_summary']['chemicals']:
            ws4.cell(row=row, column=1, value='Chemical')
            ws4.cell(row=row, column=2, value=d['name'])
            ws4.cell(row=row, column=3, value=d['total_quantity']).number_format = '#,##0.00'
            ws4.cell(row=row, column=4, value=d['unit'])
            ws4.cell(row=row, column=5, value='')
            ws4.cell(row=row, column=6, value=d['incident_count'])
            if row % 2 == 0:
                for col in range(1, 7):
                    ws4.cell(row=row, column=col).fill = alt_fill
            row += 1
        for d in report_data['damage_summary']['apparatus']:
            ws4.cell(row=row, column=1, value='Apparatus')
            ws4.cell(row=row, column=2, value=d['name'])
            ws4.cell(row=row, column=3, value=d['total_quantity'])
            ws4.cell(row=row, column=4, value=d['unit'])
            ws4.cell(row=row, column=5, value='')
            ws4.cell(row=row, column=6, value=d['incident_count'])
            if row % 2 == 0:
                for col in range(1, 7):
                    ws4.cell(row=row, column=col).fill = alt_fill
            row += 1
        self._auto_width(ws4, 6)

        # Sheet 5 - Current Stock
        ws5 = wb.create_sheet('Current Stock')
        ws5.sheet_properties.tabColor = 'F59E0B'
        headers = ['Type', 'Name', 'Current Qty', 'Unit', 'Reorder Level', 'Status']
        for col, h in enumerate(headers, 1):
            cell = ws5.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        row = 2
        for c in report_data['current_stock']['low_stock_chemicals']:
            ws5.cell(row=row, column=1, value='Chemical')
            ws5.cell(row=row, column=2, value=c['name'])
            ws5.cell(row=row, column=3, value=c['current_quantity']).number_format = '#,##0.00'
            ws5.cell(row=row, column=4, value=c['unit'])
            ws5.cell(row=row, column=5, value=c['reorder_level']).number_format = '#,##0.00'
            ws5.cell(row=row, column=6, value='Low')
            if row % 2 == 0:
                for col in range(1, 7):
                    ws5.cell(row=row, column=col).fill = alt_fill
            row += 1
        for a in report_data['current_stock']['low_stock_apparatus']:
            ws5.cell(row=row, column=1, value='Apparatus')
            ws5.cell(row=row, column=2, value=a['name'])
            ws5.cell(row=row, column=3, value=a['current_quantity'])
            ws5.cell(row=row, column=4, value=a['unit'])
            ws5.cell(row=row, column=5, value=a['reorder_level'])
            ws5.cell(row=row, column=6, value='Low')
            if row % 2 == 0:
                for col in range(1, 7):
                    ws5.cell(row=row, column=col).fill = alt_fill
            row += 1
        self._auto_width(ws5, 6)

        # Sheet 6 - Restock Recommendations
        ws6 = wb.create_sheet('Restock Recommendations')
        ws6.sheet_properties.tabColor = '8B5CF6'
        headers = ['Type', 'Name', 'Avg Monthly Usage', 'Projected Annual Need',
                   'Current Stock', 'Recommended Purchase', 'Unit']
        for col, h in enumerate(headers, 1):
            cell = ws6.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        row = 2
        for r in report_data['restock_recommendations']:
            ws6.cell(row=row, column=1, value=r['type'])
            ws6.cell(row=row, column=2, value=r['name'])
            ws6.cell(row=row, column=3, value=r['avg_monthly_usage']).number_format = '#,##0.00'
            ws6.cell(row=row, column=4, value=r['projected_annual_need']).number_format = '#,##0.00'
            ws6.cell(row=row, column=5, value=r['current_stock']).number_format = '#,##0.00'
            ws6.cell(row=row, column=6, value=r['recommended_purchase']).number_format = '#,##0.00'
            ws6.cell(row=row, column=7, value=r['unit'])
            if row % 2 == 0:
                for col in range(1, 8):
                    ws6.cell(row=row, column=col).fill = alt_fill
            row += 1
        self._auto_width(ws6, 7)

        wb.save(response)
        return response

    def _auto_width(self, ws, max_col):
        for col in range(1, max_col + 1):
            letter = get_column_letter(col)
            max_len = 15
            for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, min(len(str(cell.value)) + 2, 40))
            ws.column_dimensions[letter].width = max_len
